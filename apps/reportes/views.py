import io
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Sum, Count
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.clientes.models import Cliente
from apps.pagos.models import Pago
from apps.pedidos.models import Pedido
from apps.servicios.models import DetalleServicio
from core.permissions import EsEmpleadoActivo


class AnalyticsView(APIView):
    permission_classes = [EsEmpleadoActivo]

    def get(self, request):
        hoy = timezone.localdate()
        hace_7 = hoy - timedelta(days=6)
        hace_6m = hoy.replace(day=1) - timedelta(days=1)
        hace_6m = (hace_6m.replace(day=1) - timedelta(days=1)).replace(day=1)
        inicio_mes = hoy.replace(day=1)

        # Ingresos últimos 7 días
        ingresos_semana_qs = (
            Pago.objects.filter(
                fecha_pago__date__gte=hace_7,
                estado_pago=Pago.PAGADO,
            )
            .annotate(dia=TruncDate("fecha_pago"))
            .values("dia")
            .annotate(total=Sum("monto"))
            .order_by("dia")
        )
        dias_map = {r["dia"]: float(r["total"]) for r in ingresos_semana_qs}
        ingresos_semana = []
        for i in range(7):
            d = hace_7 + timedelta(days=i)
            ingresos_semana.append({
                "fecha": d.strftime("%d %b"),
                "total": dias_map.get(d, 0),
            })

        # Ingresos últimos 6 meses
        inicio_6m = (hoy.replace(day=1) - timedelta(days=150)).replace(day=1)
        ingresos_mes_qs = (
            Pago.objects.filter(
                fecha_pago__date__gte=inicio_6m,
                estado_pago=Pago.PAGADO,
            )
            .annotate(mes=TruncMonth("fecha_pago"))
            .values("mes")
            .annotate(total=Sum("monto"))
            .order_by("mes")
        )
        ingresos_mes = [
            {"mes": r["mes"].strftime("%b %Y"), "total": float(r["total"])}
            for r in ingresos_mes_qs
        ]

        # Pedidos por estado
        estados_qs = (
            Pedido.objects.values("estado")
            .annotate(count=Count("id"))
            .order_by("estado")
        )
        ESTADO_LABELS = {
            "pendiente":  "Pendiente",
            "en_proceso": "En proceso",
            "listo":      "Listo",
            "entregado":  "Entregado",
            "cancelado":  "Cancelado",
        }
        pedidos_por_estado = [
            {"estado": ESTADO_LABELS.get(r["estado"], r["estado"]), "count": r["count"]}
            for r in estados_qs
        ]

        # Top 5 servicios más solicitados
        top_servicios_qs = (
            DetalleServicio.objects.select_related("id_servicio")
            .values("id_servicio__nombre_servicio")
            .annotate(count=Count("id"))
            .order_by("-count")[:5]
        )
        top_servicios = [
            {"nombre": r["id_servicio__nombre_servicio"], "count": r["count"]}
            for r in top_servicios_qs
        ]

        # Resumen general
        pedidos_hoy = Pedido.objects.filter(fecha_ingreso__date=hoy).count()
        ingresos_hoy = Pago.objects.filter(
            fecha_pago__date=hoy, estado_pago=Pago.PAGADO
        ).aggregate(t=Sum("monto"))["t"] or Decimal("0")
        pedidos_mes = Pedido.objects.filter(fecha_ingreso__date__gte=inicio_mes).count()
        clientes_total = Cliente.objects.count()

        return Response({
            "ingresos_semana":    ingresos_semana,
            "ingresos_mes":       ingresos_mes,
            "pedidos_por_estado": pedidos_por_estado,
            "top_servicios":      top_servicios,
            "resumen": {
                "pedidos_hoy":    pedidos_hoy,
                "ingresos_hoy":   float(ingresos_hoy),
                "pedidos_mes":    pedidos_mes,
                "clientes_total": clientes_total,
            },
        })


class ExcelPedidosView(APIView):
    permission_classes = [EsEmpleadoActivo]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos"

        # Cabecera
        headers = ["Código", "Cliente", "Empleado", "Estado", "Fecha ingreso", "Fecha entrega", "Total (S/)"]
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        ESTADO_LABELS = {
            "pendiente":  "Pendiente",
            "en_proceso": "En proceso",
            "listo":      "Listo",
            "entregado":  "Entregado",
            "cancelado":  "Cancelado",
        }

        pedidos = Pedido.objects.select_related("id_cliente", "id_empleado").order_by("-fecha_ingreso")
        alt_fill = PatternFill("solid", fgColor="F1F5F9")

        for row_num, p in enumerate(pedidos, 2):
            fill = alt_fill if row_num % 2 == 0 else None
            values = [
                p.codigo,
                f"{p.id_cliente.nombres} {p.id_cliente.apellidos}",
                f"{p.id_empleado.nombres} {p.id_empleado.apellidos}",
                ESTADO_LABELS.get(p.estado, p.estado),
                p.fecha_ingreso.strftime("%d/%m/%Y %H:%M") if p.fecha_ingreso else "",
                p.fecha_entrega.strftime("%d/%m/%Y %H:%M") if p.fecha_entrega else "",
                float(p.total),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="center")

        # Ancho de columnas
        col_widths = [14, 28, 24, 14, 20, 20, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Hoja de pagos
        ws2 = wb.create_sheet("Pagos")
        headers2 = ["#", "Pedido (código)", "Monto (S/)", "Método", "Fecha", "Estado"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 22

        METODO_LABELS = {
            "efectivo": "Efectivo", "tarjeta": "Tarjeta",
            "transferencia": "Transferencia", "yape": "Yape", "plin": "Plin",
        }
        pagos = Pago.objects.select_related("id_pedido").order_by("-fecha_pago")
        for row_num, pg in enumerate(pagos, 2):
            fill = alt_fill if row_num % 2 == 0 else None
            values2 = [
                pg.id,
                pg.id_pedido.codigo,
                float(pg.monto),
                METODO_LABELS.get(pg.metodo_pago, pg.metodo_pago),
                pg.fecha_pago.strftime("%d/%m/%Y %H:%M") if pg.fecha_pago else "",
                pg.estado_pago.capitalize(),
            ]
            for col, val in enumerate(values2, 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                if fill:
                    cell.fill = fill
        for i, w in enumerate([8, 16, 14, 16, 20, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"lavanderia_jireh_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
